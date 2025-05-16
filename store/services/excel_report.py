import openpyxl
from openpyxl.utils import get_column_letter
from .report_generator import ReportGenerator
from io import BytesIO
from django.http import HttpResponse

class ExcelReportGenerator(ReportGenerator):
    def generate(self, products):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"

        headers = ["Nombre", "Precio", "Stock", "Categoría"]
        ws.append(headers)

        for product in products:
            ws.append([product.name, float(product.price), product.stock, product.category.name if product.category else ""])

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=productos.xlsx"
        return response
